#!/usr/bin/env python3
"""
Score Whiteboard Quiz responses against the answer key.

Reads:
  Whiteboard Quiz — Answer Key.xlsx   (gold answers in row 2)
  Whiteboard Quiz — Responses.xlsx    (team responses, row 2 onwards)

Writes:
  Whiteboard Quiz — Scores.xlsx       (per-whiteboard scores + ranking)

Scoring per whiteboard:
  Q1 (max 1 pt)  — unigram/bigram n-gram recall (case/punctuation ignored)
  Q2 (max 2 pts) — BERTScore F1, scaled to 2 pts
  Q3 (max 3 pts) — BERTScore F1, scaled to 3 pts
"""

import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from bert_score import score as bertscore

DIR       = Path(__file__).parent
KEY_FILE  = DIR / "Whiteboard Quiz — Answer Key.xlsx"
RESP_FILE = DIR / "Whiteboard Quiz — Responses.xlsx"
OUT_FILE  = DIR / "Whiteboard Quiz — Scores.xlsx"

WHITEBOARD_NUMBERS = [1, 7, 8, 10, 11, 12, 15, 17, 18, 21, 24]
Q_MAX     = [1, 2, 3]
MAX_TOTAL = sum(Q_MAX) * len(WHITEBOARD_NUMBERS)   # 66

BERT_MODEL = "distilbert-base-uncased"   # fast; swap for "roberta-large" for higher quality

# ---------------------------------------------------------------------------
# Q1 scoring: n-gram recall (unigram + bigram)
# ---------------------------------------------------------------------------

def tokenize(text: str) -> list:
    if not text:
        return []
    text = str(text).lower()
    text = re.sub(r"[^\w\s]", " ", text)
    return [t for t in text.split() if t]


def ngram_recall(a_tok: list, g_tok: list, n: int):
    if len(g_tok) < n:
        return None
    def make(tokens):
        return Counter(tuple(tokens[i : i + n]) for i in range(len(tokens) - n + 1))
    g_ng = make(g_tok)
    a_ng = make(a_tok)
    overlap = sum((a_ng & g_ng).values())
    total   = sum(g_ng.values())
    return overlap / total if total else 0.0


def ngram_score(answer, gold, max_pts: int) -> float:
    g_tok = tokenize(gold)
    a_tok = tokenize(answer)
    if not g_tok:
        return float(max_pts)
    if not a_tok:
        return 0.0
    recalls = [r for n in (1, 2) if (r := ngram_recall(a_tok, g_tok, n)) is not None]
    proportion = sum(recalls) / len(recalls)
    return round(proportion * max_pts, 2)


# ---------------------------------------------------------------------------
# Q2 / Q3 scoring: BERTScore F1 scaled to max points
# ---------------------------------------------------------------------------

def bertscore_batch(candidates: list, references: list, max_pts_list: list) -> list:
    """
    Score a batch of (candidate, reference) pairs with BERTScore.
    Returns a list of floats, one score per pair, scaled to the corresponding max_pts.
    Empty answers get 0; empty gold gets full marks.
    """
    scores = []
    active_idx, active_cands, active_refs, active_max = [], [], [], []

    for i, (cand, ref, mp) in enumerate(zip(candidates, references, max_pts_list)):
        cand_str = str(cand).strip() if cand else ""
        ref_str  = str(ref).strip()  if ref  else ""
        if not ref_str:
            scores.append(float(mp))      # empty gold → full marks
        elif not cand_str:
            scores.append(0.0)            # empty answer → 0
        else:
            scores.append(None)           # to be filled by BERTScore
            active_idx.append(i)
            active_cands.append(cand_str)
            active_refs.append(ref_str)
            active_max.append(mp)

    if active_cands:
        _, _, F1 = bertscore(
            active_cands, active_refs,
            model_type=BERT_MODEL,
            verbose=False,
            device=None,
        )
        for list_pos, (idx, mp) in enumerate(zip(active_idx, active_max)):
            scores[idx] = round(float(F1[list_pos]) * mp, 2)

    return scores


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------

def load_gold() -> list:
    """Return list of (q1_gold, q2_gold, q3_gold) per whiteboard."""
    ws = openpyxl.load_workbook(KEY_FILE).active
    row = [c.value for c in ws[2]]
    answers = []
    for i in range(len(WHITEBOARD_NUMBERS)):
        base = 2 + i * 3
        answers.append(tuple(row[base : base + 3]))
    return answers


def load_responses() -> list:
    ws = openpyxl.load_workbook(RESP_FILE).active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        team = row[1]
        if team and str(team).strip() not in ("", "—", "GOLD ANSWERS"):
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Score all teams
# ---------------------------------------------------------------------------

def compute_scores(gold: list, responses: list) -> list:
    n_teams = len(responses)
    n_wb    = len(WHITEBOARD_NUMBERS)

    # Collect Q2 and Q3 for a single BERTScore batch call
    q2_cands, q2_refs, q2_max = [], [], []
    q3_cands, q3_refs, q3_max = [], [], []
    q1_scores_flat = []   # computed inline (fast)

    for row in responses:
        for i in range(n_wb):
            base = 2 + i * 3
            def get(idx):
                return row[idx] if idx < len(row) else None
            a1, a2, a3 = get(base), get(base + 1), get(base + 2)
            q1_scores_flat.append(ngram_score(a1, gold[i][0], Q_MAX[0]))
            q2_cands.append(a2); q2_refs.append(gold[i][1]); q2_max.append(Q_MAX[1])
            q3_cands.append(a3); q3_refs.append(gold[i][2]); q3_max.append(Q_MAX[2])

    print(f"Running BERTScore on {len(q2_cands)} Q2 pairs …")
    q2_scores_flat = bertscore_batch(q2_cands, q2_refs, q2_max)
    print(f"Running BERTScore on {len(q3_cands)} Q3 pairs …")
    q3_scores_flat = bertscore_batch(q3_cands, q3_refs, q3_max)

    # Reassemble into per-team, per-whiteboard structure
    results = []
    for t, row in enumerate(responses):
        team_name = str(row[1]).strip()
        wb_scores = []
        for i in range(n_wb):
            flat = t * n_wb + i
            wb_scores.append([
                q1_scores_flat[flat],
                q2_scores_flat[flat],
                q3_scores_flat[flat],
            ])
        total = round(sum(s for wb in wb_scores for s in wb), 2)
        results.append({"team": team_name, "wb": wb_scores, "total": total})

    return results


# ---------------------------------------------------------------------------
# Write output workbook
# ---------------------------------------------------------------------------

BLUE   = PatternFill("solid", fgColor="4472C4")
GREEN  = PatternFill("solid", fgColor="70AD47")
GOLD   = PatternFill("solid", fgColor="FFD966")
SUBTOT = PatternFill("solid", fgColor="BDD7EE")
ROW_A  = PatternFill("solid", fgColor="FFFFFF")
ROW_B  = PatternFill("solid", fgColor="EBF3FB")
WHITE  = Font(bold=True, color="FFFFFF")
BOLD   = Font(bold=True)
CTR    = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(wrap_text=True, horizontal="center", vertical="center")


def write_detail_sheet(ws, results: list):
    # Grand total in column B, then per-whiteboard scores
    header = ["Team", f"Grand\nTotal\n/{MAX_TOTAL}"]
    for n in WHITEBOARD_NUMBERS:
        header += [f"WB{n}\nQ1\n/1", f"WB{n}\nQ2\n/2", f"WB{n}\nQ3\n/3", f"WB{n}\nTotal\n/6"]

    for col, val in enumerate(header, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.fill      = GREEN if col == 2 else BLUE
        c.font      = WHITE
        c.alignment = WRAP

    ws.row_dimensions[1].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    for col in range(3, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.freeze_panes = "C2"

    for r, result in enumerate(results):
        row_num = r + 2
        fill = ROW_A if r % 2 == 0 else ROW_B

        c = ws.cell(row=row_num, column=1, value=result["team"])
        c.font = BOLD
        c.fill = fill

        # Grand total in column B
        c = ws.cell(row=row_num, column=2, value=result["total"])
        c.font      = BOLD
        c.fill      = GOLD
        c.alignment = CTR

        col = 3
        for wb_scores in result["wb"]:
            for score in wb_scores:
                c = ws.cell(row=row_num, column=col, value=score)
                c.alignment = CTR
                c.fill = fill
                col += 1
            wb_total = round(sum(wb_scores), 2)
            c = ws.cell(row=row_num, column=col, value=wb_total)
            c.font      = BOLD
            c.fill      = SUBTOT
            c.alignment = CTR
            col += 1


def write_summary_sheet(ws, results: list):
    for col, val in enumerate(["Rank", "Team", "Total Score", f"/ {MAX_TOTAL}"], 1):
        c = ws.cell(row=1, column=col, value=val)
        c.fill      = BLUE
        c.font      = WHITE
        c.alignment = CTR

    for rank, result in enumerate(sorted(results, key=lambda x: x["total"], reverse=True), 1):
        ws.append([rank, result["team"], result["total"], MAX_TOTAL])
        for col in range(1, 5):
            ws.cell(row=rank + 1, column=col).alignment = CTR

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.row_dimensions[1].height = 20


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading answer key …")
    gold = load_gold()

    print("Loading responses …")
    responses = load_responses()
    if not responses:
        print(f"No team responses found in {RESP_FILE.name}.")
        return

    print(f"Scoring {len(responses)} team(s) across {len(WHITEBOARD_NUMBERS)} whiteboards …\n")
    results = compute_scores(gold, responses)

    wb_out = openpyxl.Workbook()
    write_detail_sheet(wb_out.active, results)
    wb_out.active.title = "Scores per Whiteboard"
    write_summary_sheet(wb_out.create_sheet("Summary"), results)
    wb_out.save(OUT_FILE)

    print(f"\nSaved: {OUT_FILE.name}")
    print(f"\n{'Team':<30} {'Score':>6} / {MAX_TOTAL}")
    print("-" * 42)
    for result in sorted(results, key=lambda x: x["total"], reverse=True):
        print(f"{result['team']:<30} {result['total']:>6.2f}")


if __name__ == "__main__":
    main()
